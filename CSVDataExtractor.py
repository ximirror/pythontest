from csv import reader as CSVReader
import sys
import openpyxl
from openpyxl import Workbook


class Column:
    
    def __init__(self, _startRow: int, _column: int):    
        self.startRow = _startRow
        self.column = _column
        self.data = []

    def setcolumn(self, columnID: str):
        columnID_lowercase = columnID.lower()
        strlength=len(columnID_lowercase)
        column_temp=0
        for index in range(strlength):
            print(f"{columnID_lowercase[index]} at {index}")
            column_temp=column_temp + (ord(columnID_lowercase[index]) - ord('a') + 1) * pow(26, index)
        self.column=column_temp - 1

    def add(self, line, currentLine):
        if currentLine >= self.startRow:
            self.data.append(float(str(line[self.column]).replace(',', '.')))
            
    def print(self):
        for _data in self.data:
            print(_data)
            
class XLSXImporter:
    
    def __init__(self, _file):
        print("__init__(XLSXImporter)")
        self.file = _file
        self.reader: Workbook = None
        
    def open(self):
        try:
            self.reader = openpyxl.load_workbook(self.file, read_only=True)
            self.importData = self.reader.active
            assert(None is not self.importData)
        except:
            print(sys.exc_info())
            raise Exception("something wrong with the imported file")
        
    def doImport(self, _columns):
        currentRow: int = 0
        for row in self.importData.iter_rows(1, self.importData.max_row, values_only=True):
            for column in _columns:
                column.add(row, currentRow)
            currentRow += 1
        return True
    
    def close(self):
        self.reader.close()
                
    def __enter__(self):
        print("__enter__(XLSXImporter)")
        if self.open():
            return self
        return None
    
    def __exit__(self, exc_type, exc_value, traceback):
        print("__exit__(XLSXImporter)")
        self.close()
            
class CSVImporter:
    
    def __init__(self, _file, _columns):
        print("__init__")
        self.file = _file
        self.columns = _columns
        self.reader: CSVReader = None
        
    def open(self):
        with open(self.file, "r") as importFile:
            self.reader = CSVReader(importFile)
            currentRow: int = 0
            for row in self.reader:
                for column in self.columns:
                    column.add(row, currentRow)
                currentRow += 1
    
    def close(self):
        None
                
    def __enter__(self):
        print("__enter__")
        self.open()
        return self
    
    def __exit__(self, exc_type, exc_value, traceback):
        print("__exit__")
        self.close()

if __name__ == "__main__":
    #importer = XLSXImporter("/home/xi/Downloads/Test1.xlsx")
    #importer.open()
    #columns = []
    #columns.append(Column(23, 0))
    #columns.append(Column(23, 4))
    #importer.doImport(columns)
    #importer.close()
    #columns[0].print()
    #columns[1].print()

    column = Column(0, 0)
    column.setcolumn("AA")
    print(f"{column.column}, shall be 26")
    column.setcolumn("AC")
    print(f"{column.column}, shall be 28")
    column.setcolumn("AF")
    print(f"{column.column}, shall be 31")