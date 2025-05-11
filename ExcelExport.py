import shutil
import sys
from typing import List
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet

import Parameter

class Column:
    
    def __init__(self, _columnNumber: int, _rowNumber: int, _data: List):
        self.column: int = _columnNumber
        self.row: int = _rowNumber
        self.data: List = _data
        
    def getData(self, _row) -> str:
        if _row < self.row:
            return None
        if _row > self.row + len(self.data) - 1:
            return None
        return self.data[_row - self.row]
    
    def getColumn(self) -> int:
        return self.column
    
class Averager:
    
    def average(_data: List) -> float:
        avg: float = 0
        for f in _data:
            avg += f
        avg /= len(_data)
        return avg
    
    def __init__(self, _columnNumber: int, _rowNumber: int, _data: List):
        self.base: Column = Column(_columnNumber = _columnNumber, _rowNumber=_rowNumber, _data=[Averager.average(_data)])
        
    def getData(self, _row) -> str:
        return self.base.getData(_row)
    
    def getColumn(self) -> int:
        return self.base.getColumn()
        
    
class XLSXWriter:
    
    def __init__(self, _parameter: Parameter):
        self.parameter: Parameter = _parameter
        self.output = None
        self.template = None
        
    def open(self):
        try:
            shutil.copyfile(self.parameter.templatePath, self.parameter.outputPath)
            # self.template = openpyxl.load_workbook(self.parameter.templatePath, rich_text=True, read_only=True, keep_vba=True)
            # self.templateData = self.template[self.parameter.templateSheet]
            # self.output = Workbook(write_only=True)
            self.output = openpyxl.load_workbook(self.parameter.outputPath, rich_text=True, read_only=False, keep_vba=True)
            #assert(self.templateData is not None)
            assert(self.output is not None)
        except:
            print(repr(sys.exc_info()))
            raise Exception("something happened while opening the template or creating the output file")
        
    def doExport(self, _columns: List):
        try:
            #sheet: WriteOnlyWorksheet = self.output.create_sheet()
            sheet: WriteOnlyWorksheet = self.output[self.parameter.templateSheet]
            assert(self.output is not None)

            rowNr: int = 0
            
            #for row in self.templateData.iter_rows(1, self.templateData.max_row):
            for row in sheet.iter_rows(1, sheet.max_row):
                for importDataColumn in _columns:
                    val = importDataColumn.getData(rowNr)
                    if val is not None:
                        row[importDataColumn.getColumn()].value = val
                
                rowNr += 1
                #sheet.append(row)
        except:
            print("at index row {}".format(rowNr))
            print(repr(sys.exc_info()))
            raise Exception("something wrong while exporting data")

        
    def write(self, _outputPath: str):
        assert(self.output is not None)
        self.output.save(_outputPath)
    
    def close(self):
        #self.template.close()
        self.output.close()
                
    def __enter__(self):
        print("__enter__(XLSXWriter)")
        if self.open():
            return self
        return None
    
    def __exit__(self, exc_type, exc_value, traceback):
        print("__exit__(XLSXWriter)")
        self.close()
        